from typing import List
from openpyxl import load_workbook


def main(filepath, line_range: List[int]):
    wb = load_workbook(filename=filepath)
    ws = wb.active  # Get the first sheet

    # Iterate over all rows in the worksheet for the lines specified by line_range    
    if line_range is None:
        start = 1
        end = ws.max_row
    else:
        start, end = line_range

    for row in ws.iter_rows(min_row=start+1, max_row=end, values_only=True):
        print(row)

def parse_range(range_str):
    if range_str is None:
        return None
    if "-" in range_str:
        line_range = range_str.split("-")
    elif ":" in range_str:
        line_range = range_str.split(":")
    
    if len(line_range) == 1:
        return [0, int(line_range[0])]
    elif len(line_range) == 2:
        return [int(line_range[0]), int(line_range[1])]
    else:
        raise ValueError(f"Invalid line range: {range_str}")

def parse_args():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("filepath", type=str, help="path to the xlsx file")
    parser.add_argument("--lines", type=parse_range, default=None, help="range of lines to display, such as '10' or '0-10'")
    args = parser.parse_args()
    return args

if __name__ == '__main__':    
    args = parse_args()
    main(args.filepath, args.lines)